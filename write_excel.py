import requests
import time
from openpyxl import load_workbook
from config import BASIC_API

def get_data(api_name: str, district_id):
    r = requests.get(f"{BASIC_API}{api_name}?district_id={district_id}")
    r_data = r.json()
    data = r_data['data']
    return data

def excel_bus(template_filename: str, api_name: str, district_id):
    data = get_data(api_name, district_id)
    wb = load_workbook(f"templates/{template_filename}.xlsx")
    sheet = wb.active
    row, col = 3, 1
    for item in data:
        sheet.cell(row, col).value = item['rownum']
        sheet.cell(row, col + 1).value = item['object_name']
        sheet.cell(row, col + 2).value = item['org_tin']
        sheet.cell(row, col + 3).value = item['mobile_phone']
        row += 1
    at_time = int(time.time())
    wb.save(f"Тадбиркорлик субектлари{at_time}.xlsx")
    return at_time

def excel_women(template_filename: str, api_name: str, district_id):
    data = get_data(api_name, district_id)
    wb = load_workbook(f"templates/{template_filename}.xlsx")
    sheet = wb.active
    row, col = 3, 1
    for item in data:
        sheet.cell(row, col).value = item['rownum']
        sheet.cell(row, col + 1).value = item['full_name']
        sheet.cell(row, col + 2).value = item['passport_number']
        sheet.cell(row, col + 3).value = item['phone']
        sheet.cell(row, col + 4).value = item['direction_name']
        row += 1
    at_time = int(time.time())
    wb.save(f"Аёллар дафтари{at_time}.xlsx")
    return at_time

# def excel_young(template_filename: str, api_name: str, district_id):
#     data = get_data(api_name, district_id)
#     wb = load_workbook(f"templates/{template_filename}.xlsx")
#     sheet = wb.active
#     row, col = 3, 1
#     for item in data:
#         sheet.cell(row, col).value = item['rownum']
#         sheet.cell(row, col + 1).value = item['full_name']
#         sheet.cell(row, col + 2).value = item['passport_number']
#         sheet.cell(row, col + 3).value = item['phone']
#         sheet.cell(row, col + 4).value = item['direction_name']
#         row += 1
#     at_time = int(time.time())
#     wb.save(f"Yoshlar_daftari{at_time}.xlsx")
#     return at_time

def excel_iron(template_filename: str, api_name: str, district_id):
    data = get_data(api_name, district_id)
    wb = load_workbook(f"templates/{template_filename}.xlsx")
    sheet = wb.active
    row, col = 3, 1
    for item in data:
        sheet.cell(row, col).value = item['rownum']
        sheet.cell(row, col + 1).value = item['full_name']
        sheet.cell(row, col + 2).value = item['pinfl']
        sheet.cell(row, col + 3).value = item['phone']
        sheet.cell(row, col + 4).value = item['status_value']
        sheet.cell(row, col + 5).value = item['registrationdate']
        sheet.cell(row, col + 6).value = item['liquidationdate']
        row += 1
    at_time = int(time.time())
    wb.save(f"Темир дафтар{at_time}.xlsx")
    return at_time

def excel_social(template_filename: str, api_name: str, district_id):
    data = get_data(api_name, district_id)
    wb = load_workbook(f"templates/{template_filename}.xlsx")
    sheet = wb.active
    row, col = 3, 1
    for item in data:
        sheet.cell(row, col).value = item['rownum']
        sheet.cell(row, col + 1).value = item['name']
        sheet.cell(row, col + 2).value = item['birth_date']
        sheet.cell(row, col + 3).value = item['pinfl']
        sheet.cell(row, col + 4).value = item['phone']
        sheet.cell(row, col + 5).value = item['payment_start_date']
        sheet.cell(row, col + 6).value = item['payment_end_date']
        row += 1
    at_time = int(time.time())
    wb.save(f"Ягона ижтимоий реестр{at_time}.xlsx")
    return at_time
